"""
Excel大日程 → ScheduleData JSON 変換スクリプト

L4開発大日程（全体共有）.xlsx のセルデータを解析し、
ScheduleData v3.3.0 形式のJSONに変換する。

対象シート:
  - 大日程 (事業開発) → ページ「事業開発」
  - 大日程(技術開発)  → ページ「技術開発」
"""

import json
import os
import sys
import re
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl")
    sys.exit(1)


# ── 設定 ──

EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "L4開発大日程（全体共有）.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "schedule", "l4_schedule.json")

# タイムライン設定
TIMELINE_START = "2025-10"
TIMELINE_END = "2032-12"
MONTH_WIDTH_PX = 30
LANE_HEADER_WIDTH_PX = 160

# バーの高さ
BAR_HEIGHT_PX = 22
BAR_SPACING = 28  # Y方向の最小間隔


# ── 色判定 ──

COLOR_KEYWORDS = {
    "blue": ["実証", "実験", "テスト", "試験", "検証", "トライ", "評価", "DC", "走行"],
    "green": ["計画", "準備", "設計", "方針", "検討", "仕様", "開発", "構築", "推進", "トレーニング", "教育"],
    "orange": ["認可", "認証", "許可", "許認可", "Fix", "SOP"],
    "pink": ["契約", "営業", "顧客", "商用", "サービス", "運用", "合意", "交渉"],
    "purple": ["生産", "量産", "架装", "工事", "工程"],
    "red": ["★", "マイルストーン"],
}


def classify_color(text):
    """セル値のキーワードから色を判定"""
    if not text:
        return "gray"
    text_str = str(text).strip()
    for color, keywords in COLOR_KEYWORDS.items():
        for kw in keywords:
            if kw in text_str:
                return color
    return "gray"


# ── 列→月マッピング ──

def build_column_to_month_map(ws, year_row, month_row):
    """ヘッダー行から列番号→"YYYY-MM"の辞書を構築"""
    col_map = {}
    current_year = None

    for col in range(4, ws.max_column + 1):
        year_val = ws.cell(row=year_row, column=col).value
        if year_val is not None:
            current_year = int(year_val)

        month_val = ws.cell(row=month_row, column=col).value
        if month_val is not None and current_year is not None:
            month_int = int(month_val)
            col_map[col] = f"{current_year:04d}-{month_int:02d}"

    return col_map


# ── バー生成ロジック ──

def extract_bars_from_row(ws, row_num, col_map, label, bar_id_prefix, start_idx):
    """行のD列以降を走査し、連続セルをバーに変換"""
    bars = []
    sorted_cols = sorted(col_map.keys())

    # 連続セル群を見つける
    current_run = []  # [(col, value), ...]
    for col in sorted_cols:
        cell_val = ws.cell(row=row_num, column=col).value
        if cell_val is not None:
            cell_str = str(cell_val).strip()
            if cell_str:
                current_run.append((col, cell_str))
                continue
        # セルが空 → 現在のランを確定
        if current_run:
            bars.append(_make_bar(current_run, col_map, label, bar_id_prefix, start_idx + len(bars)))
            current_run = []

    # 最後のランを確定
    if current_run:
        bars.append(_make_bar(current_run, col_map, label, bar_id_prefix, start_idx + len(bars)))

    return bars


def _make_bar(run, col_map, label, prefix, idx):
    """連続セルのランからScheduleBarを生成"""
    start_col = run[0][0]
    end_col = run[-1][0]
    cell_values = [v for _, v in run]

    # 色をセル値から判定（最初のセルの値を優先）
    color = classify_color(cell_values[0])

    # ユニークなセル値を集めてtooltipに
    unique_vals = []
    for v in cell_values:
        v_clean = v.strip()
        if v_clean and v_clean not in unique_vals:
            unique_vals.append(v_clean)
    tooltip = "→".join(unique_vals) if unique_vals else None

    # ラベルが長すぎる場合は短縮
    bar_label = label
    if len(bar_label) > 30:
        bar_label = bar_label[:28] + "…"

    return {
        "id": f"{prefix}_{idx}",
        "label": bar_label,
        "startMonth": col_map[start_col],
        "endMonth": col_map[end_col],
        "color": color,
        "yOffsetInLane": 0,  # 後で計算
        "heightPx": BAR_HEIGHT_PX,
        "tooltip": tooltip,
    }


# ── Y座標の自動計算 ──

def month_to_int(month_str):
    """'YYYY-MM' を整数（YYYY*12+MM）に変換"""
    parts = month_str.split("-")
    return int(parts[0]) * 12 + int(parts[1])


def assign_y_offsets(bars):
    """バーのyOffsetInLaneを自動計算（重なり回避）"""
    if not bars:
        return 0

    # 各バーの占有期間
    intervals = []
    for bar in bars:
        start = month_to_int(bar["startMonth"])
        end = month_to_int(bar["endMonth"])
        intervals.append((start, end, bar))

    # 開始月順にソート
    intervals.sort(key=lambda x: x[0])

    # グリーディにY位置を割り当て
    rows = []  # [(end_month, y_offset), ...]
    for start, end, bar in intervals:
        # 空いている最小のYを探す
        assigned_y = None
        for i, (row_end, row_y) in enumerate(rows):
            if start > row_end:  # この行は空いている
                bar["yOffsetInLane"] = row_y
                rows[i] = (end, row_y)
                assigned_y = row_y
                break
        if assigned_y is None:
            new_y = len(rows) * BAR_SPACING + 4
            bar["yOffsetInLane"] = new_y
            rows.append((end, new_y))

    max_y = max(bar["yOffsetInLane"] for bar in bars)
    return max_y + BAR_HEIGHT_PX + 8  # 必要な最小高さ


# ── 事業開発シート処理 ──

# C列にこれらの値がある行はカテゴリヘッダー（レーン）ではなく日付情報
BIZ_DATE_PATTERNS = re.compile(r"^\d{2}/\d{1,2}$|^本番開始$|^\d{2}年度")


def _is_biz_category(a_str, c_val):
    """事業開発シートのカテゴリ行（レーンヘッダー）かどうかを判定
    条件: A列が通常テキスト（・/　で始まらない）かつ C列に担当者名がある"""
    if a_str.startswith("・") or a_str.startswith("\u3000") or a_str.startswith("　"):
        return False
    c_str = str(c_val).strip() if c_val else ""
    if c_str and not BIZ_DATE_PATTERNS.match(c_str):
        return True
    return False


def process_biz_sheet(ws, col_map):
    """事業開発シートを解析してswimLanes配列を返す"""
    swim_lanes = []
    current_lane = None
    last_task_label = ""
    bar_counter = 0

    for row in range(9, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        c_val = ws.cell(row=row, column=3).value

        # A列が空の行: 前のタスクの継続として扱う
        if a_val is None:
            has_data = any(ws.cell(row=row, column=col).value is not None for col in col_map)
            if has_data and current_lane is not None:
                cont_label = last_task_label if last_task_label else "(続き)"
                bars = extract_bars_from_row(ws, row, col_map, cont_label, f"bar_biz", bar_counter)
                current_lane["bars"].extend(bars)
                bar_counter += len(bars)
            continue

        a_str = str(a_val).strip()
        if not a_str:
            continue

        # D列以降にデータがあるか
        has_data = any(ws.cell(row=row, column=col).value is not None for col in col_map)
        person = str(c_val).strip() if c_val else ""

        if _is_biz_category(a_str, c_val):
            # カテゴリ行 → 新しいSwimLane
            current_lane = _new_biz_lane(a_str, person, len(swim_lanes))
            swim_lanes.append(current_lane)
            last_task_label = a_str

            if has_data:
                bars = extract_bars_from_row(ws, row, col_map, a_str, f"bar_biz", bar_counter)
                _add_person_tooltip(bars, person)
                current_lane["bars"].extend(bars)
                bar_counter += len(bars)

        elif current_lane is not None:
            # タスク行（・付き, 　付き, or 通常テキスト）→ 現在のレーンにバー追加
            task_label = a_str.lstrip("・").lstrip("\u3000").lstrip("　").strip()
            last_task_label = task_label
            if has_data:
                bars = extract_bars_from_row(ws, row, col_map, task_label, f"bar_biz", bar_counter)
                _add_person_tooltip(bars, person)
                current_lane["bars"].extend(bars)
                bar_counter += len(bars)

        else:
            # カテゴリでもなく、親レーンもない場合: 独立レーンとして扱う
            last_task_label = a_str
            if has_data:
                current_lane = _new_biz_lane(a_str, "", len(swim_lanes))
                swim_lanes.append(current_lane)
                bars = extract_bars_from_row(ws, row, col_map, a_str, f"bar_biz", bar_counter)
                current_lane["bars"].extend(bars)
                bar_counter += len(bars)

    # 空レーンを除去
    swim_lanes = [lane for lane in swim_lanes if lane["bars"] or lane["milestones"]]

    # Y座標計算とレーン高さ調整
    for lane in swim_lanes:
        min_height = assign_y_offsets(lane["bars"])
        lane["heightPx"] = max(60, min_height)

    return swim_lanes


def _new_biz_lane(label, person, idx):
    """新しいSwimLaneを作成"""
    lane_label = label
    if person:
        lane_label = f"{label}\n({person})"
    return {
        "id": f"lane_biz_{idx}",
        "label": lane_label,
        "heightPx": 80,
        "bars": [],
        "milestones": [],
        "tags": [],
        "registryId": f"tmpl_biz_{idx}",
    }


def _add_person_tooltip(bars, person):
    """バーのtooltipに担当者情報を追加"""
    if not person or BIZ_DATE_PATTERNS.match(person):
        return
    for b in bars:
        b["tooltip"] = f"[{person}] " + (b.get("tooltip") or "")


# ── 技術開発シート処理 ──

# 技術開発シートのセクション定義（B列のカテゴリ行ベース）
TECH_SECTIONS = [
    # (label, start_row, end_row)
    ("イベント\n事業計画", 3, 13),
    ("AD機能開発\n基本機能", 14, 39),
    ("AD機能開発\nODD拡張", 40, 44),
    ("車両\nGen1", 45, 58),
    ("Gen2.x\nSWPF開発", 59, 66),
    ("Gen2.x\nHWPF開発", 67, 95),
    ("認可・許可", 96, 103),
    ("Gen3\n量産", 104, 122),
    ("生産準備", 123, 134),
    ("ToS結合\nシステム開発", 135, 162),
    ("生産計画", 163, 175),
]


def process_tech_sheet(ws, col_map):
    """技術開発シートを解析してswimLanes配列を返す"""
    swim_lanes = []
    bar_counter = 0

    # D-F列のマージセル情報を収集（カテゴリラベル用、スケジュールデータから除外）
    merged_label_cols = set()  # (row, col) のセット
    for merged in ws.merged_cells.ranges:
        if merged.min_col >= 4 and merged.min_col <= 6:
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_label_cols.add((r, c))

    # スケジュールデータ用の列マップ（マージされたカテゴリ列を除外）
    def get_schedule_col_map(row):
        return {col: month for col, month in col_map.items()
                if (row, col) not in merged_label_cols}

    for sec_idx, (sec_label, start_row, end_row) in enumerate(TECH_SECTIONS):
        lane = {
            "id": f"lane_tech_{sec_idx}",
            "label": sec_label,
            "heightPx": 80,
            "bars": [],
            "milestones": [],
            "tags": [],
            "registryId": f"tmpl_tech_{sec_idx}",
        }

        for row in range(start_row, end_row + 1):
            row_col_map = get_schedule_col_map(row)

            # G列(col=7)以降にスケジュールデータがあるか確認
            # D-F列(col=4-6)はカテゴリ/車両ラベル列なのでスケジュールデータとしない
            schedule_col_map = {col: month for col, month in row_col_map.items() if col >= 7}
            has_schedule_data = False
            for col in schedule_col_map:
                v = ws.cell(row=row, column=col).value
                if v is not None and str(v).strip():
                    has_schedule_data = True
                    break

            if not has_schedule_data:
                continue

            # ラベル決定: B列 > C列 > E列（車両名等）
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            e_val = ws.cell(row=row, column=5).value

            label = ""
            if b_val:
                label = str(b_val).strip()
            elif c_val:
                label = str(c_val).strip()
            elif e_val and str(e_val).strip():
                e_str = str(e_val).strip()
                if "使用" not in e_str and "車両" not in e_str and "シャーシ" not in e_str:
                    label = e_str

            if not label:
                # A/B/C/E列にラベルがない場合、最初のセル値をラベルに使う
                for col in sorted(schedule_col_map.keys()):
                    v = ws.cell(row=row, column=col).value
                    if v is not None and str(v).strip():
                        first_val = str(v).strip()
                        if len(first_val) > 25:
                            first_val = first_val[:23] + "…"
                        label = first_val
                        break
                if not label:
                    label = f"(行{row})"

            bars = extract_bars_from_row(ws, row, schedule_col_map, label, f"bar_tech", bar_counter)
            lane["bars"].extend(bars)
            bar_counter += len(bars)

        if lane["bars"] or lane["milestones"]:
            min_height = assign_y_offsets(lane["bars"])
            lane["heightPx"] = max(60, min_height)
            swim_lanes.append(lane)

    return swim_lanes


# ── メイン処理 ──

def build_schedule_data(biz_lanes, tech_lanes):
    """ScheduleData JSONを構築"""
    # laneRegistry 生成
    lane_registry = []
    all_lanes = biz_lanes + tech_lanes
    for lane in all_lanes:
        lane_registry.append({
            "id": lane["registryId"],
            "label": lane["label"],
            "tags": lane.get("tags", []),
            "defaultHeightPx": lane["heightPx"],
        })

    data = {
        "version": "3.3.0",
        "lastModified": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z"),
        "timeline": {
            "startDate": TIMELINE_START,
            "endDate": TIMELINE_END,
            "monthWidthPx": MONTH_WIDTH_PX,
            "laneHeaderWidthPx": LANE_HEADER_WIDTH_PX,
        },
        "pages": [
            {
                "id": "p_biz",
                "name": "事業開発",
                "swimLanes": biz_lanes,
                "annotations": [],
                "connections": [],
            },
            {
                "id": "p_tech",
                "name": "技術開発",
                "swimLanes": tech_lanes,
                "annotations": [],
                "connections": [],
            },
        ],
        "laneRegistry": lane_registry,
        "settings": {
            "fontSizeLaneTitle": 11,
            "fontSizeBarText": 10,
            "fontSizeMilestone": 10,
            "fontSizeCalendar": 10,
            "zoomLevel": "month",
            "displayMode": "fit",
            "showTooltips": True,
            "showMemos": False,
            "themeMode": "light",
        },
    }
    return data


def main():
    print(f"Excel: {EXCEL_PATH}")
    print(f"Output: {OUTPUT_PATH}")

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    # Excel読み込み
    print("\n[1] Excelファイルを読み込み中...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── 事業開発シート ──
    print("\n[2] 事業開発シートを処理中...")
    ws_biz = wb["大日程 (事業開発)"]
    col_map_biz = build_column_to_month_map(ws_biz, year_row=7, month_row=8)
    print(f"   列マッピング: {len(col_map_biz)}列 ({min(col_map_biz.values())} 〜 {max(col_map_biz.values())})")

    biz_lanes = process_biz_sheet(ws_biz, col_map_biz)
    biz_bars = sum(len(l["bars"]) for l in biz_lanes)
    print(f"   レーン数: {len(biz_lanes)}, バー数: {biz_bars}")
    for lane in biz_lanes:
        print(f"     [{lane['label'].replace(chr(10), ' ')}] バー={len(lane['bars'])}, 高さ={lane['heightPx']}px")

    # ── 技術開発シート ──
    print("\n[3] 技術開発シートを処理中...")
    ws_tech = wb["大日程(技術開発)"]
    col_map_tech = build_column_to_month_map(ws_tech, year_row=1, month_row=2)
    print(f"   列マッピング: {len(col_map_tech)}列 ({min(col_map_tech.values())} 〜 {max(col_map_tech.values())})")

    tech_lanes = process_tech_sheet(ws_tech, col_map_tech)
    tech_bars = sum(len(l["bars"]) for l in tech_lanes)
    print(f"   レーン数: {len(tech_lanes)}, バー数: {tech_bars}")
    for lane in tech_lanes:
        print(f"     [{lane['label'].replace(chr(10), ' ')}] バー={len(lane['bars'])}, 高さ={lane['heightPx']}px")

    # ── JSON構築・出力 ──
    print("\n[4] JSON構築中...")
    data = build_schedule_data(biz_lanes, tech_lanes)

    # 出力ディレクトリ確認
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"   保存完了: {OUTPUT_PATH}")
    total_bars = biz_bars + tech_bars
    total_lanes = len(biz_lanes) + len(tech_lanes)
    print(f"\n完了! レーン={total_lanes}, バー={total_bars}")


if __name__ == "__main__":
    main()
